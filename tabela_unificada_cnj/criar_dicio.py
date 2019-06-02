import openpyxl


def converter_tabelas(nome_da_tabela):
    dict_save = open(nome_da_tabela[:-5] + '.py', mode='w', encoding='UTF8')
    wb = openpyxl.load_workbook(nome_da_tabela)
    ws = wb['Planilha1']

    assuntos = []
    codigos = []
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, max_col=5):
        for cell in row:
            if cell.value is not None:
                assuntos.append(cell.value)
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, min_col=6, max_col=6):
        for cell in row:
            if cell.value is not None:
                codigos.append(str(cell.value))

    dicionario = dict(zip(assuntos, codigos))
    dict_save.writelines('assuntos_cnj = ' + str(dicionario))
    dict_save.close()


converter_tabelas(r'14_Tabela_Assuntos_CNJ.xlsx')
